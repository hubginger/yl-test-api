# coding=utf-8
"""
    openpyxl 读取 Excel 文件
    file_name = Delivery_System_V1.5.xlsx

    11 个 sheet 页, 176 条数据, 每个 sheet 都根据配置文件找 index 比读取依次慢 0.5 s

"""
import json
import os
import openpyxl

from typing import Dict
from openpyxl.utils.exceptions import InvalidFileException

from common.utils.do_conf import do_conf
from common.utils.do_log import yl_log
from common.utils.do_path import DATA_FOLDER


# @Time    :  2023-12-29 17:36:11
# @Author  :  jiangtong
# @Email   :  jiangtong@yljt.cn
# @Project :  yljk_test_api
# @File    :  do_excel


class ExcelData:
    """
        用例数据对象
        将 Excel 数据加载出来之后, 将表头和数据聚合成该对象, 表头为属性, 数据为属性值
    """

    # 假的类属性, 只是为了顺利 . 出代码
    # ExcelData.case_id
    url, method, host, = None, None, 'a',
    expected_response: dict = None
    case_id, title, = None, None,
    sheet_name, row = None, None,
    data: dict = None

    def __init__(self, zip_obj):
        """
            zip_obj , zip 嵌套 tuple
            tuple 统一, 长度为 2
            tuple[0], Excel 表头
            tuple[1], Excel 数据
            将 tuple[0] 和 tuple[1] 聚合对象的属性和值

              - case_id
              - title
              - host
              - url
              - method
              - data
              - expected_response
        """
        # 通过 obj. 时, 有代码提示
        # self.case_id, self.title, self.url, self.method, self.data, self.expected, self.expected_response = None, None, None, None, None, None, None
        for obj in zip_obj:
            setattr(self, obj[0], obj[1])

    def __str__(self):
        """
            方便获取对象的属性和值, 生效于直接打印时
        """
        return str(self.__dict__)

    def __repr__(self):
        """
            方便获取对象的属性和值, 生效于返回传递时
        """
        return str(self.__dict__)


class DoExcel:
    """
        Excel 操作类
        其中 __get_column_s 和 __get_column 不建议每次用例执行时都调用
        建议 __get_column_s 和 __get_column 的结果放在配置文件中
        show 方法用于调试, 比如修改了 Excel 列, 可以先来运行 show 方法
        __json 是为了将 '{' 开头的数据直接处理成字典

        read
            读取 Excel 的某一 Sheet
            将数据聚合成 ExcelData 对象, 使用时方便调用, excel_data_obj.case_id

        read_all
            读取 Excel 的所有 Sheet
            内部逻辑是拿到所有 Sheet, 然后循环调用 read 方法
    """

    def __init__(self, file_name, sheet_name=None):
        self.file_name = os.sep.join([DATA_FOLDER, file_name])
        self.sheet_name = sheet_name

    @staticmethod
    def __xls_to_xlsx(old_file):
        """
            openpyxl 打不开 xls 文件
            这里将 xls 后缀处理成 xlsx
            有些场景下修改后会导致无法打开 :
                因只是系统层面改后缀, 如果是太低版本的 MicroSoft Excel 保存为 xls 文件
                则需要使用较高版本 MicroSoft Excel 打开 Excel 文档另存为 xlsx 文件
        """
        new_file = old_file.replace('.xls', '.xlsx')
        os.rename(old_file, new_file)
        return new_file

    def __get_column(self, sheet_name=None):
        """
            获取回写列的索引
        """

        # xls 转 xlsx
        try:
            wb = openpyxl.load_workbook(self.file_name)
        except InvalidFileException:
            wb = openpyxl.load_workbook(self.__xls_to_xlsx(self.file_name))

        # sheet_name, sheet 对象, 列索引, 配置的表头, Excel 回写列
        sheet_name = sheet_name if sheet_name else self.sheet_name
        sheet_stream = wb[sheet_name]
        list_head = [cell.value for cell in sheet_stream[2]]
        write_back = do_conf.read_one('excel', 'write_back')

        # 返回回写列的索引
        for index, head in enumerate(list_head):
            if head == write_back:
                return index + 1

    def __get_column_s(self, sheet_name=None):
        """
            根据 excel.yml 中配置的表头 ( column ), 查询 Excel 对应的列
            将这些列的索引放到列表中返回
        """

        # xls 转 xlsx
        try:
            wb = openpyxl.load_workbook(self.file_name)
        except InvalidFileException:
            wb = openpyxl.load_workbook(self.__xls_to_xlsx(self.file_name))

        # sheet_name, sheet 对象, 列索引, 配置的表头, Excel 表头数据
        sheet_name = sheet_name if sheet_name else self.sheet_name
        sheet_stream, res_index_s, = wb[sheet_name], [],
        headers, datas = do_conf.read_one('excel', 'column'), [cell.value for cell in sheet_stream[2]]

        # 索引
        for index, head in enumerate(datas):
            if head in headers:
                res_index_s.append(index + 1)
        wb.close()
        return res_index_s

    def read(self):
        """
            读取 Excel 的某个 Sheet 页的所有数据
            返回格式为:
                {
                    'sheet_name':{
                        '用例编号_1': row_1: obj,
                        '用例编号_2': row_2: obj,
                        '用例编号_3': row_3: obj,
                    },
                }
            格式说明:
                快速根据 '用例编号' 匹配到 row
                基础格式和 read_all 保持一致, 区别在于 read 返回只有一个键值对, read_all 是多个键值对
        """
        # xls 转 xlsx 再加载
        try:
            wb = openpyxl.load_workbook(self.file_name)
        except InvalidFileException:
            wb = openpyxl.load_workbook(self.__xls_to_xlsx(self.file_name))

        # sheet 对象, 表头值, 表头索引, 返回
        sheet_stream = wb[self.sheet_name]
        list_head_use = do_conf.read_one('excel', 'column')
        column_index_s = do_conf.read_one('excel', 'index')
        datas, _num = dict(), 3
        _res: Dict[str, Dict[str, ExcelData]] = dict()

        # 读取数据
        for row in sheet_stream.iter_rows(min_row=3, values_only=True):
            if row:
                list_row_use = [self.__json(row[_index - 1]) for _index in column_index_s]
                case_data = ExcelData(zip(list_head_use, list_row_use)) if row and row[0] else None
                if case_data:
                    case_data.sheet_name = self.sheet_name
                    case_data.row = _num
                    _num += 1
                    datas[row[0]] = case_data
        _res[self.sheet_name] = datas
        wb.close()
        return _res

    def read_all(self, _only_sheet: str = None, _skip_sheet_s: list = None):
        """

            读取 Excel 的所有数据

            _only_sheet :    指定仅读取某 sheet , 字符串
            _skip_sheet_s :  指定忽略读取哪些 sheet , 列表
            _only_sheet 的优先级高于 _skip_sheet_s, 指定读取某 Sheet , 就仅读取该 sheet, 不管 _skip_sheet_s 中是否指定忽略该 sheet

            返回格式:
                {
                    'sheet_1':{
                        '用例编号_1': row_1: obj,
                        '用例编号_2': row_2: obj,
                        '用例编号_3': row_3: obj,
                    },
                    'sheet_2':{
                        '用例编号_1': row_1: obj,
                        '用例编号_2': row_2: obj,
                        '用例编号_3': row_3: obj,
                    },
                    ...
                }
        """
        # 加载 Excel 当后缀为 xls 时转为 xlsx 再加载
        sheet_s = []
        try:
            wb = openpyxl.load_workbook(self.file_name)
        except InvalidFileException:
            wb = openpyxl.load_workbook(self.__xls_to_xlsx(self.file_name))
        except FileNotFoundError:
            yl_log.exception(f'Excel文件: << {self.file_name} >> 不存在, 请检查文件名或路径.')
            return dict()
        else:
            # _only_sheet 和 _skip_sheet_s, 其中 _only_sheet 的优先级高于 _skip_sheet_s
            if _only_sheet:
                sheet_s = [_only_sheet]
                yl_log.warning(f'读取 Excel, 注意当前仅读取 Sheet: << {_only_sheet} >>')
            else:
                skip_s = ['备注', '说明', '声明', '注意', '示例', 'demo']
                skip_s.extend(_skip_sheet_s) if _skip_sheet_s else None
                sheet_s = [sheet for sheet in wb.sheetnames if sheet not in skip_s]
        wb.close()

        # 读取所有 sheet, sheet_name 为键, sheet 页所有数据为值
        _res: Dict[str, Dict[str, ExcelData]] = dict()
        for sheet in sheet_s:
            self.sheet_name = sheet
            try:
                sheet_data = self.read()
                _res.update(sheet_data)
            except (KeyError,):
                yl_log.exception(f'读取 Excel 的 Sheet 页: <<{sheet}>> 出现异常, 继续读取下一个 sheet 页 ')
                continue
        return _res

    def show(self):
        """
        输出 :
            Excel 回写列的索引
            Excel 被参数化的列的索引
            Excel 所有的 Sheet 页名称
        """
        print('Excel 回写列的索引: ')
        print('     ', self.__get_column())
        print()

        print('Excel 被参数化的列的索引: ')
        print('     ', self.__get_column_s())
        print()

        print('以下打印可直接复制粘贴到 excel.yml 中')
        for index in self.__get_column_s():
            print(f'  - {index}')
        print()

        try:
            wb = openpyxl.load_workbook(self.file_name)
        except InvalidFileException:
            wb = openpyxl.load_workbook(self.__xls_to_xlsx(self.file_name))
        print('Excel 的所有 Sheet 页:')
        print('     ', wb.sheetnames)
        print('     ', len(wb.sheetnames))
        wb.close()

    @staticmethod
    def __json(target: str):
        """
            转 json 成功, 则返回 json
            转 json 失败, 则返回 target

            target 类型非 str, ---> TypeError
            target 内容非 json, ---> JSONDecodeError
        """
        _res = target
        try:
            if target and (target.strip().startswith('{') or target.strip().startswith('[')):
                _res = json.loads(target)
        except (TypeError, json.decoder.JSONDecodeError,):
            pass
        return _res


if __name__ == '__main__':
    # __get_column / __get_column_s
    # """
    DoExcel('Delivery_System_V1.5.xlsx', '登录模块').show()
    # """

    # read
    """
    do_excel = DoExcel('Delivery_System_V1.5.xlsx', '登录模块')
    print('datas:', do_excel.read())
    # """

    # read_all
    """
    import time

    begin = time.time()
    do_excel = DoExcel('Delivery_System_V1.5.xlsx')
    _all = do_excel.read_all()
    _nums = 0
    # sheet_name :
    for k, v in _all.items():
        # case_id :
        for _k, _v in v.items():
            _nums += 1
            print(f'{k}: {_k}, {_v}')
    end = time.time()
    duration = end - begin
    print('读取条数:', _nums)
    print('读取耗时:', duration)
    # """
